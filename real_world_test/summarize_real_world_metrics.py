#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import math
import time
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from statistics import mean, pstdev
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl


DEFAULT_TOPK_PERCENTAGES = [5, 10, 20, 50]


@dataclass
class RunMetrics:
    controller_run_id: str
    project_name: str
    variant: str
    repeat_index: int
    result_dir: Path
    review_excel: Path
    metrics_json: Path
    benchmark_bug_total: int
    tp_known: int
    fn: int
    recall_known: float
    raw_report_count: int
    issue_count: int
    reduction_ratio: float
    non_benchmark_issue_count: int
    inspection_burden: float
    topk_recall: Dict[int, float]
    topk_k: Dict[int, int]
    mrr: float
    non_benchmark_issue_signatures: set[str]
    pipeline_total_sec: float
    soot_facts_generation_sec: float
    scan_total_sec: float
    input_tokens: int
    output_tokens: int
    total_tokens: int
    query_count: int
    non_benchmark_issue_jaccard_mean: Optional[float] = None


def _load_json(path: Path) -> Dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _resolve_repo_local_path(raw_path: object) -> Path:
    text = _normalize_text(raw_path)
    path = Path(text)
    if path.exists():
        return path.resolve()

    normalized = text.replace("\\", "/")
    for marker in ["/result/", "/real_world_test/"]:
        idx = normalized.find(marker)
        if idx != -1:
            relative = normalized[idx + 1 :]
            candidate = (Path(__file__).resolve().parents[1] / relative).resolve()
            if candidate.exists() or marker == "/result/":
                return candidate
    return path.resolve()


def _parse_json_cell(value: object) -> List[str]:
    text = str(value or "").strip()
    if text == "":
        return []
    try:
        data = json.loads(text)
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    return [str(item) for item in data]


def _load_sheet_rows(workbook_path: Path, sheet_name: str) -> List[Dict[str, object]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {workbook_path}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) == 0:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: List[Dict[str, object]] = []
    for row in rows[1:]:
        item = {}
        for idx, header in enumerate(headers):
            item[header] = row[idx] if idx < len(row) else None
        result.append(item)
    return result


def _normalize_text(value: object) -> str:
    return str(value or "").strip()


def _safe_int(value: object, default: int = 0) -> int:
    try:
        text = str(value).strip()
        if text == "":
            return default
        return int(float(text))
    except Exception:
        return default


def _safe_float(value: object, default: float = 0.0) -> float:
    try:
        text = str(value).strip()
        if text == "":
            return default
        return float(text)
    except Exception:
        return default


def _jaccard(a: set[str], b: set[str]) -> float:
    if not a and not b:
        return 1.0
    return len(a & b) / len(a | b)


def _issue_signature(entry: Dict[str, object]) -> str:
    metadata = entry.get("metadata", {})
    if not isinstance(metadata, dict):
        metadata = {}
    parts = [
        _normalize_text(metadata.get("source_file")).lower().replace("\\", "/"),
        _normalize_text(metadata.get("issue_primary_defect_method_uid")),
        _normalize_text(metadata.get("source_method_uid")),
        _normalize_text(metadata.get("leak_root_method_uid")),
        _normalize_text(metadata.get("source_line")),
        _normalize_text(metadata.get("resource_kind")),
        _normalize_text(metadata.get("release_context")),
        _normalize_text(metadata.get("guarantee_level")),
        _normalize_text(metadata.get("source_symbol")),
        ",".join(sorted(_normalize_text(x) for x in metadata.get("issue_component_keys", []) if x is not None))
        if isinstance(metadata.get("issue_component_keys"), list)
        else "",
        _normalize_text(entry.get("buggy_value")),
    ]
    return "||".join(parts)


def _compute_topk_metrics(
    detect_order: Sequence[str],
    benchmark_hits_rows: Sequence[Dict[str, object]],
    topk_percentages: Sequence[int],
) -> Tuple[int, int, float, Dict[int, float], Dict[int, int], float, set[str]]:
    total_detect = len(detect_order)
    position_by_report = {str(report_id): idx + 1 for idx, report_id in enumerate(detect_order)}

    benchmark_bug_total = 0
    tp_known = 0
    fn = 0
    reciprocal_ranks: List[float] = []
    first_ranks: List[Optional[int]] = []
    matched_report_ids_union: set[str] = set()

    for row in benchmark_hits_rows:
        benchmark_bug_total += 1
        final_label = _normalize_text(row.get("final_label"))
        matched_ids = _parse_json_cell(row.get("matched_report_ids"))
        matched_report_ids_union.update(matched_ids)
        ranks = [position_by_report[rid] for rid in matched_ids if rid in position_by_report]
        first_rank = min(ranks) if ranks else None
        first_ranks.append(first_rank)
        if final_label == "TP_known":
            tp_known += 1
        elif final_label == "FN":
            fn += 1
        reciprocal_ranks.append(1.0 / first_rank if first_rank is not None else 0.0)

    recall_known = tp_known / benchmark_bug_total if benchmark_bug_total > 0 else 0.0
    mrr = sum(reciprocal_ranks) / benchmark_bug_total if benchmark_bug_total > 0 else 0.0

    topk_recall: Dict[int, float] = {}
    topk_k: Dict[int, int] = {}
    for pct in topk_percentages:
        k = max(1, math.ceil(total_detect * (pct / 100.0))) if total_detect > 0 else 0
        hits = sum(
            1 for rank in first_ranks if rank is not None and rank <= k
        )
        topk_k[pct] = k
        topk_recall[pct] = hits / benchmark_bug_total if benchmark_bug_total > 0 else 0.0

    return (
        benchmark_bug_total,
        tp_known,
        recall_known,
        topk_recall,
        topk_k,
        mrr,
        matched_report_ids_union,
    )


def _load_run_metrics(run_row: Dict[str, object], topk_percentages: Sequence[int]) -> RunMetrics:
    result_dir = _resolve_repo_local_path(run_row.get("result_dir"))
    review_excel = result_dir / "review_units.xlsx"
    metrics_json = result_dir / "run_metrics_raw.json"
    detect_info_json = result_dir / "detect_info.json"
    detect_info_raw_json = result_dir / "detect_info_raw.json"
    detect_issue_stats_json = result_dir / "detect_info_issue_stats.json"

    detect_info = _load_json(detect_info_json)
    detect_order = [str(key) for key in detect_info.keys()]
    detect_issue_stats = _load_json(detect_issue_stats_json) if detect_issue_stats_json.exists() else {}
    benchmark_hits_rows = _load_sheet_rows(review_excel, "benchmark_hits")
    metrics_payload = _load_json(metrics_json)

    (
        benchmark_bug_total,
        tp_known,
        recall_known,
        topk_recall,
        topk_k,
        mrr,
        matched_report_ids_union,
    ) = _compute_topk_metrics(detect_order, benchmark_hits_rows, topk_percentages)

    fn = benchmark_bug_total - tp_known
    raw_report_count = _safe_int(
        detect_issue_stats.get("raw_report_count"),
        default=len(_load_json(detect_info_raw_json)) if detect_info_raw_json.exists() else len(detect_info),
    )
    issue_count = _safe_int(detect_issue_stats.get("issue_count"), default=len(detect_info))
    reduction_ratio = _safe_float(
        detect_issue_stats.get("reduction_ratio"),
        default=((raw_report_count - issue_count) / raw_report_count if raw_report_count > 0 else 0.0),
    )

    non_benchmark_issue_signatures: set[str] = set()
    for report_id, entry in detect_info.items():
        if str(report_id) in matched_report_ids_union:
            continue
        if isinstance(entry, dict):
            non_benchmark_issue_signatures.add(_issue_signature(entry))
    non_benchmark_issue_count = len(non_benchmark_issue_signatures)
    inspection_burden = (
        non_benchmark_issue_count / benchmark_bug_total if benchmark_bug_total > 0 else 0.0
    )

    timing = metrics_payload.get("timing", {})
    llm_usage = metrics_payload.get("llm_usage", {})

    return RunMetrics(
        controller_run_id=_normalize_text(run_row.get("controller_run_id")),
        project_name=_normalize_text(run_row.get("project_name")),
        variant=_normalize_text(run_row.get("variant")),
        repeat_index=_safe_int(run_row.get("repeat_index")),
        result_dir=result_dir,
        review_excel=review_excel,
        metrics_json=metrics_json,
        benchmark_bug_total=benchmark_bug_total,
        tp_known=tp_known,
        fn=fn,
        recall_known=recall_known,
        raw_report_count=raw_report_count,
        issue_count=issue_count,
        reduction_ratio=reduction_ratio,
        non_benchmark_issue_count=non_benchmark_issue_count,
        inspection_burden=inspection_burden,
        topk_recall=topk_recall,
        topk_k=topk_k,
        mrr=mrr,
        non_benchmark_issue_signatures=non_benchmark_issue_signatures,
        pipeline_total_sec=_safe_float(timing.get("pipeline_total_sec")),
        soot_facts_generation_sec=_safe_float(timing.get("soot_facts_generation_sec")),
        scan_total_sec=_safe_float(timing.get("scan_total_sec")),
        input_tokens=_safe_int(llm_usage.get("input_tokens")),
        output_tokens=_safe_int(llm_usage.get("output_tokens")),
        total_tokens=_safe_int(llm_usage.get("total_tokens")),
        query_count=_safe_int(llm_usage.get("query_count")),
    )


def _load_ledger_rows(experiment_run_dir: Path) -> List[Dict[str, object]]:
    ledger_path = experiment_run_dir / "experiment_runs.json"
    data = _load_json(ledger_path)
    runs = data.get("runs", [])
    if not isinstance(runs, list):
        raise ValueError(f"Invalid runs payload in {ledger_path}")
    result = []
    for row in runs:
        if not isinstance(row, dict):
            continue
        if _normalize_text(row.get("variant")) == "__soot_generation__":
            continue
        if _normalize_text(row.get("status")) != "success":
            continue
        if _normalize_text(row.get("result_dir")) == "":
            continue
        result.append(row)
    return result


def _load_result_dir_row(result_dir: Path) -> Dict[str, object]:
    meta_path = result_dir / "experiment_run_meta.json"
    if meta_path.exists():
        meta = _load_json(meta_path)
    else:
        meta = {}

    project_name = _normalize_text(meta.get("project_name")) or result_dir.parent.name
    controller_run_id = _normalize_text(meta.get("controller_run_id")) or f"imported_{project_name}"
    variant = _normalize_text(meta.get("variant")) or "unknown"
    repeat_index = _safe_int(meta.get("repeat_index"), default=0)

    return {
        "controller_run_id": controller_run_id,
        "project_name": project_name,
        "variant": variant,
        "repeat_index": repeat_index,
        "result_dir": str(result_dir.resolve()),
        "status": "success",
    }


def _collect_result_rows_from_roots(result_roots: Sequence[Path]) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for root in result_roots:
        if not root.exists():
            raise FileNotFoundError(f"Result root not found: {root}")
        for child in sorted(root.iterdir()):
            if not child.is_dir():
                continue
            if not (child / "review_units.xlsx").exists():
                continue
            if not (child / "run_metrics_raw.json").exists():
                continue
            if not (child / "detect_info.json").exists():
                continue
            rows.append(_load_result_dir_row(child))
    rows.sort(key=lambda row: (_normalize_text(row.get("project_name")), _normalize_text(row.get("variant")), _safe_int(row.get("repeat_index")), _normalize_text(row.get("result_dir"))))
    return rows


def _attach_nonbenchmark_jaccard(run_metrics: List[RunMetrics]) -> None:
    grouped: Dict[Tuple[str, str], List[RunMetrics]] = defaultdict(list)
    for item in run_metrics:
        grouped[(item.project_name, item.variant)].append(item)

    for _key, items in grouped.items():
        for item in items:
            scores = [
                _jaccard(item.non_benchmark_issue_signatures, other.non_benchmark_issue_signatures)
                for other in items
                if other is not item
            ]
            item.non_benchmark_issue_jaccard_mean = mean(scores) if scores else 1.0


def _run_metrics_to_row(item: RunMetrics, topk_percentages: Sequence[int]) -> Dict[str, object]:
    row = {
        "controller_run_id": item.controller_run_id,
        "project_name": item.project_name,
        "variant": item.variant,
        "repeat_index": item.repeat_index,
        "result_dir": str(item.result_dir),
        "benchmark_bug_total": item.benchmark_bug_total,
        "tp_known": item.tp_known,
        "fn": item.fn,
        "benchmark_hit_rate": item.recall_known,
        "raw_report_count": item.raw_report_count,
        "issue_count": item.issue_count,
        "reduction_ratio": item.reduction_ratio,
        "non_benchmark_issue_count": item.non_benchmark_issue_count,
        "inspection_burden": item.inspection_burden,
        "mrr": item.mrr,
        "non_benchmark_issue_jaccard_mean": item.non_benchmark_issue_jaccard_mean,
        "pipeline_total_sec": item.pipeline_total_sec,
        "soot_facts_generation_sec": item.soot_facts_generation_sec,
        "scan_total_sec": item.scan_total_sec,
        "input_tokens": item.input_tokens,
        "output_tokens": item.output_tokens,
        "total_tokens": item.total_tokens,
        "query_count": item.query_count,
    }
    for pct in topk_percentages:
        row[f"top_{pct}pct_k"] = item.topk_k[pct]
        row[f"top_{pct}pct_recall"] = item.topk_recall[pct]
    return row


def _aggregate_group(
    items: Sequence[RunMetrics], topk_percentages: Sequence[int]
) -> Dict[str, object]:
    if not items:
        return {}

    def series(getter):
        return [getter(item) for item in items]

    summary = {
        "project_name": items[0].project_name,
        "variant": items[0].variant,
        "repeat_count": len(items),
        "benchmark_bug_total": items[0].benchmark_bug_total,
        "tp_known_mean": mean(series(lambda x: x.tp_known)),
        "fn_mean": mean(series(lambda x: x.fn)),
        "benchmark_hit_rate_mean": mean(series(lambda x: x.recall_known)),
        "raw_report_count_mean": mean(series(lambda x: x.raw_report_count)),
        "issue_count_mean": mean(series(lambda x: x.issue_count)),
        "reduction_ratio_mean": mean(series(lambda x: x.reduction_ratio)),
        "non_benchmark_issue_count_mean": mean(series(lambda x: x.non_benchmark_issue_count)),
        "inspection_burden_mean": mean(series(lambda x: x.inspection_burden)),
        "mrr_mean": mean(series(lambda x: x.mrr)),
        "non_benchmark_issue_jaccard_mean": mean(
            series(lambda x: x.non_benchmark_issue_jaccard_mean or 0.0)
        ),
        "pipeline_total_sec_mean": mean(series(lambda x: x.pipeline_total_sec)),
        "scan_total_sec_mean": mean(series(lambda x: x.scan_total_sec)),
        "total_tokens_mean": mean(series(lambda x: x.total_tokens)),
    }

    if len(items) > 1:
        summary["benchmark_hit_rate_std"] = pstdev(series(lambda x: x.recall_known))
        summary["issue_count_std"] = pstdev(series(lambda x: x.issue_count))
        summary["mrr_std"] = pstdev(series(lambda x: x.mrr))
        summary["pipeline_total_sec_std"] = pstdev(series(lambda x: x.pipeline_total_sec))
        summary["total_tokens_std"] = pstdev(series(lambda x: x.total_tokens))
    else:
        summary["benchmark_hit_rate_std"] = 0.0
        summary["issue_count_std"] = 0.0
        summary["mrr_std"] = 0.0
        summary["pipeline_total_sec_std"] = 0.0
        summary["total_tokens_std"] = 0.0

    for pct in topk_percentages:
        values = series(lambda x: x.topk_recall[pct])
        summary[f"top_{pct}pct_recall_mean"] = mean(values)
        summary[f"top_{pct}pct_recall_std"] = pstdev(values) if len(values) > 1 else 0.0

    return summary


def summarize_real_world_metrics(
    experiment_run_dir: Optional[Path],
    topk_percentages: Sequence[int],
    ledger_rows_override: Optional[Sequence[Dict[str, object]]] = None,
) -> Tuple[List[RunMetrics], List[Dict[str, object]]]:
    if ledger_rows_override is not None:
        ledger_rows = [dict(row) for row in ledger_rows_override]
    else:
        if experiment_run_dir is None:
            raise ValueError("experiment_run_dir is required when ledger_rows_override is not provided")
        ledger_rows = _load_ledger_rows(experiment_run_dir)
    run_metrics = [_load_run_metrics(row, topk_percentages) for row in ledger_rows]
    _attach_nonbenchmark_jaccard(run_metrics)

    grouped: Dict[Tuple[str, str], List[RunMetrics]] = defaultdict(list)
    for item in run_metrics:
        grouped[(item.project_name, item.variant)].append(item)

    summary_rows = [
        _aggregate_group(sorted(items, key=lambda x: x.repeat_index), topk_percentages)
        for _key, items in sorted(grouped.items())
    ]
    return run_metrics, summary_rows


def _write_csv(path: Path, rows: Sequence[Dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([])
        return
    headers = list(rows[0].keys())
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _write_metrics_outputs(
    output_dir: Path,
    run_metrics: Sequence[RunMetrics],
    summary_rows: Sequence[Dict[str, object]],
    topk_percentages: Sequence[int],
) -> None:
    per_run_rows = [_run_metrics_to_row(item, topk_percentages) for item in run_metrics]

    per_run_json = output_dir / "real_world_metrics_per_run.json"
    per_run_csv = output_dir / "real_world_metrics_per_run.csv"
    summary_json = output_dir / "real_world_metrics_summary.json"
    summary_csv = output_dir / "real_world_metrics_summary.csv"

    output_dir.mkdir(parents=True, exist_ok=True)
    with open(per_run_json, "w", encoding="utf-8") as f:
        json.dump({"runs": per_run_rows}, f, indent=2, ensure_ascii=False)
    _write_csv(per_run_csv, per_run_rows)

    with open(summary_json, "w", encoding="utf-8") as f:
        json.dump({"summary": list(summary_rows)}, f, indent=2, ensure_ascii=False)
    _write_csv(summary_csv, summary_rows)


def _write_single_result_output(
    result_dir: Path,
    run_metric: RunMetrics,
    topk_percentages: Sequence[int],
    output_json: Optional[Path] = None,
) -> Path:
    output_path = output_json if output_json is not None else (result_dir / "real_world_metrics.json")
    payload = {
        "schema_version": "1.0",
        "result_dir": str(result_dir.resolve()),
        "metrics": _run_metrics_to_row(run_metric, topk_percentages),
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Summarize real-world RepoAudit metrics from experiment_runs."
    )
    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument(
        "--result-dir",
        help="One RepoAudit result directory; writes real_world_metrics.json beside detect_info.json by default",
    )
    input_group.add_argument(
        "--experiment-run-dir",
        help="Path to real_world_test/experiment_runs/<run_id>",
    )
    input_group.add_argument(
        "--result-root",
        action="append",
        default=[],
        help="Existing result root such as result/dfbscan/<model>/MLK/Java/fitnesse; can be repeated",
    )
    parser.add_argument(
        "--output-dir",
        help="Output directory for metrics files. Defaults to the experiment-run-dir or a generated imported directory.",
    )
    parser.add_argument(
        "--output-json",
        help="Optional output json path for --result-dir mode (default: <result-dir>/real_world_metrics.json)",
    )
    parser.add_argument(
        "--topk-percentages",
        default="5,10,20,50",
        help="Comma-separated top-k%% values, e.g. 5,10,20,50",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    topk_percentages = [int(x.strip()) for x in args.topk_percentages.split(",") if x.strip() != ""]
    if args.result_dir:
        result_dir = _resolve_repo_local_path(args.result_dir)
        run_metric = _load_run_metrics(_load_result_dir_row(result_dir), topk_percentages)
        output_path = _write_single_result_output(
            result_dir,
            run_metric,
            topk_percentages,
            Path(args.output_json).resolve() if args.output_json else None,
        )
        print(f"[OK] real-world metrics written to: {output_path}")
        return
    if args.experiment_run_dir:
        experiment_run_dir = Path(args.experiment_run_dir).resolve()
        output_dir = Path(args.output_dir).resolve() if args.output_dir else experiment_run_dir
        run_metrics, summary_rows = summarize_real_world_metrics(experiment_run_dir, topk_percentages)
    else:
        result_roots = [_resolve_repo_local_path(path) for path in args.result_root]
        ledger_rows = _collect_result_rows_from_roots(result_roots)
        if args.output_dir:
            output_dir = Path(args.output_dir).resolve()
        else:
            suffix = time.strftime("%Y%m%d-%H%M%S", time.localtime())
            output_dir = (Path(__file__).resolve().parent / "experiment_runs" / f"imported_real_world_metrics_{suffix}").resolve()
        run_metrics, summary_rows = summarize_real_world_metrics(
            None,
            topk_percentages,
            ledger_rows_override=ledger_rows,
        )

    _write_metrics_outputs(output_dir, run_metrics, summary_rows, topk_percentages)

    per_run_json = output_dir / "real_world_metrics_per_run.json"
    per_run_csv = output_dir / "real_world_metrics_per_run.csv"
    summary_json = output_dir / "real_world_metrics_summary.json"
    summary_csv = output_dir / "real_world_metrics_summary.csv"
    print(f"[OK] per-run metrics written to: {per_run_json}")
    print(f"[OK] per-run csv written to: {per_run_csv}")
    print(f"[OK] summary metrics written to: {summary_json}")
    print(f"[OK] summary csv written to: {summary_csv}")


if __name__ == "__main__":
    main()
