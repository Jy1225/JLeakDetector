#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import openpyxl


DEFAULT_WORKBOOK_NAME = "review_units.xlsx"


def _normalize_label(value: object) -> str:
    text = str(value or "").strip()
    if text == "":
        return ""
    return text


def _load_sheet_rows(workbook_path: Path, sheet_name: str) -> List[Dict[str, object]]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {workbook_path}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) == 0:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: List[Dict[str, object]] = []
    for row in rows[1:]:
        if row is None:
            continue
        item: Dict[str, object] = {}
        for idx, header in enumerate(headers):
            item[header] = row[idx] if idx < len(row) else None
        result.append(item)
    return result


def _safe_ratio(numerator: int, denominator: int) -> Optional[float]:
    if denominator <= 0:
        return None
    return numerator / denominator


def _build_summary(workbook_path: Path) -> Dict[str, object]:
    benchmark_rows = _load_sheet_rows(workbook_path, "benchmark_hits")
    review_rows = _load_sheet_rows(workbook_path, "review_units")

    tp_known = 0
    fn = 0
    benchmark_uncertain = 0
    for row in benchmark_rows:
        final_label = _normalize_label(row.get("final_label"))
        if final_label == "TP_known":
            tp_known += 1
        elif final_label == "FN":
            fn += 1
        elif final_label == "UNCERTAIN":
            benchmark_uncertain += 1

    tp_new = 0
    fp = 0
    review_uncertain = 0
    duplicate = 0
    unlabeled = 0
    done_count = 0
    todo_count = 0
    for row in review_rows:
        review_status = _normalize_label(row.get("review_status"))
        human_label = _normalize_label(row.get("human_label"))
        if review_status.lower() == "done":
            done_count += 1
        elif review_status.lower() == "todo":
            todo_count += 1

        if human_label == "TP_new":
            tp_new += 1
        elif human_label == "FP":
            fp += 1
        elif human_label == "UNCERTAIN":
            review_uncertain += 1
        elif human_label == "DUPLICATE":
            duplicate += 1
        elif human_label == "":
            unlabeled += 1

    total_tp = tp_known + tp_new
    precision = _safe_ratio(total_tp, total_tp + fp)
    recall = _safe_ratio(tp_known, tp_known + fn)
    f1 = (
        None
        if precision is None or recall is None or (precision + recall) == 0
        else (2 * precision * recall) / (precision + recall)
    )

    return {
        "schema_version": "1.0",
        "workbook_path": str(workbook_path.resolve()),
        "result_dir": str(workbook_path.resolve().parent),
        "benchmark_hits": {
            "total_rows": len(benchmark_rows),
            "tp_known": tp_known,
            "fn": fn,
            "uncertain": benchmark_uncertain,
        },
        "review_units": {
            "total_rows": len(review_rows),
            "tp_new": tp_new,
            "fp": fp,
            "uncertain": review_uncertain,
            "duplicate": duplicate,
            "unlabeled": unlabeled,
            "review_status_done": done_count,
            "review_status_todo": todo_count,
        },
        "metrics": {
            "tp_known": tp_known,
            "tp_new": tp_new,
            "fp": fp,
            "fn": fn,
            "tn": None,
            "precision": precision,
            "recall": recall,
            "f1": f1,
        },
        "notes": {
            "tn_status": "not_applicable_from_current_review_workbook",
            "precision_formula": "(TP_known + TP_new) / (TP_known + TP_new + FP)",
            "recall_formula": "TP_known / (TP_known + FN)",
        },
    }


def _write_summary_json(summary: Dict[str, object], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=4, ensure_ascii=False)


def _write_summary_csv(summary: Dict[str, object], output_path: Path) -> None:
    rows = [
        ("tp_known", summary["metrics"]["tp_known"]),
        ("tp_new", summary["metrics"]["tp_new"]),
        ("fp", summary["metrics"]["fp"]),
        ("fn", summary["metrics"]["fn"]),
        ("tn", summary["metrics"]["tn"]),
        ("precision", summary["metrics"]["precision"]),
        ("recall", summary["metrics"]["recall"]),
        ("f1", summary["metrics"]["f1"]),
        ("benchmark_total_rows", summary["benchmark_hits"]["total_rows"]),
        ("benchmark_uncertain", summary["benchmark_hits"]["uncertain"]),
        ("review_total_rows", summary["review_units"]["total_rows"]),
        ("review_uncertain", summary["review_units"]["uncertain"]),
        ("review_duplicate", summary["review_units"]["duplicate"]),
        ("review_unlabeled", summary["review_units"]["unlabeled"]),
        ("review_status_done", summary["review_units"]["review_status_done"]),
        ("review_status_todo", summary["review_units"]["review_status_todo"]),
    ]
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["metric", "value"])
        writer.writerows(rows)


def summarize_review_workbook(
    workbook_path: Path,
    json_output: Optional[Path] = None,
    csv_output: Optional[Path] = None,
) -> Dict[str, object]:
    workbook_path = workbook_path.resolve()
    summary = _build_summary(workbook_path)
    if json_output is None:
        json_output = workbook_path.parent / "review_summary.json"
    if csv_output is None:
        csv_output = workbook_path.parent / "review_summary.csv"

    _write_summary_json(summary, json_output.resolve())
    _write_summary_csv(summary, csv_output.resolve())
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Summarize manual labels from review_units.xlsx."
    )
    parser.add_argument(
        "--review-xlsx",
        default="",
        help="Path to review_units.xlsx (default: <result-dir>/review_units.xlsx)",
    )
    parser.add_argument(
        "--result-dir",
        default="",
        help="Result directory containing review_units.xlsx",
    )
    parser.add_argument(
        "--json-output",
        default="",
        help="Optional path to output review_summary.json",
    )
    parser.add_argument(
        "--csv-output",
        default="",
        help="Optional path to output review_summary.csv",
    )
    args = parser.parse_args()

    if str(args.review_xlsx).strip() != "":
        workbook_path = Path(args.review_xlsx)
    elif str(args.result_dir).strip() != "":
        workbook_path = Path(args.result_dir) / DEFAULT_WORKBOOK_NAME
    else:
        raise ValueError("Please provide --review-xlsx or --result-dir")

    json_output = (
        Path(args.json_output) if str(args.json_output).strip() != "" else None
    )
    csv_output = (
        Path(args.csv_output) if str(args.csv_output).strip() != "" else None
    )

    summary = summarize_review_workbook(
        workbook_path=workbook_path,
        json_output=json_output,
        csv_output=csv_output,
    )
    print(
        "[OK] review summary written to:",
        workbook_path.resolve().parent / "review_summary.json",
        "and",
        workbook_path.resolve().parent / "review_summary.csv",
    )
    print(
        json.dumps(
            {
                "tp_known": summary["metrics"]["tp_known"],
                "tp_new": summary["metrics"]["tp_new"],
                "fp": summary["metrics"]["fp"],
                "fn": summary["metrics"]["fn"],
                "precision": summary["metrics"]["precision"],
                "recall": summary["metrics"]["recall"],
                "f1": summary["metrics"]["f1"],
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
