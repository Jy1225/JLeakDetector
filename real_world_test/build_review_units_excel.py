#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font


DEFAULT_BENCHMARK_XLSX = Path("real_world_test/selected_real_world_projects.xlsx")


@dataclass
class BenchmarkBug:
    benchmark_bug_id: str
    project_full_name: str
    project_short_name: str
    bug_type: str
    benchmark_file: str
    benchmark_method: str
    start_line: Optional[int]
    end_line: Optional[int]
    key_variable_lines: List[int]


@dataclass
class ReportEntry:
    report_id: str
    project_name: str
    bug_type: str
    model_name: str
    result_dir: str
    primary_file: str
    primary_method: str
    source_line: Optional[int]
    source_lines: List[int]
    issue_component_keys: List[str]
    resource_kind: str
    release_context: str
    guarantee_level: str
    grouped_report_ids: List[str]
    grouped_report_count: int
    representative_buggy_value: str
    representative_explanation: str
    source_methods: List[str]


def _load_json(path: Path) -> Dict:
    if not path.exists():
        raise FileNotFoundError(f"Required file not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _safe_int(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    try:
        text = str(value).strip()
        if text == "":
            return None
        return int(float(text))
    except Exception:
        return None


def _normalize_path(path_str: str) -> str:
    return str(path_str).replace("\\", "/").strip().lower()


def _json_cell(value: object) -> str:
    return json.dumps(value, ensure_ascii=False)


def _parse_result_dir_context(result_dir: Path) -> Dict[str, str]:
    parts = result_dir.resolve().parts
    if len(parts) < 6:
        raise ValueError(f"Unexpected result dir layout: {result_dir}")
    return {
        "scanner": parts[-6],
        "model_name": parts[-5],
        "bug_type": parts[-4],
        "language": parts[-3],
        "project_name": parts[-2],
        "run_id": parts[-1],
    }


def _iter_detail_rows(ws) -> Iterable[Dict[str, object]]:
    rows = ws.iter_rows(values_only=True)
    headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    for row in rows:
        if row is None:
            continue
        item = {}
        for idx, header in enumerate(headers):
            item[header] = row[idx] if idx < len(row) else None
        yield item


def _parse_defect_method(raw: object) -> Tuple[str, str]:
    text = str(raw or "").strip()
    if text == "":
        return "", ""
    if ":" not in text:
        return text, ""
    file_part, method_part = text.rsplit(":", 1)
    return str(file_part).strip(), str(method_part).strip()


def _parse_key_variable_lines(raw: object) -> List[int]:
    text = str(raw or "").strip()
    if text == "" or text.lower() == "none":
        return []
    line_tokens: List[str] = []
    if ":" in text:
        line_tokens = [token.strip() for token in text.rsplit(":", 1)[-1].split(",")]
    else:
        match = re.search(r"\.java:?(\d+(?:,\d+)*)$", text, flags=re.IGNORECASE)
        if match is not None:
            line_tokens = [token.strip() for token in match.group(1).split(",")]
    result: List[int] = []
    for token in line_tokens:
        line_num = _safe_int(token)
        if line_num is not None:
            result.append(line_num)
    return sorted(set(result))


def _project_short_name(project_full_name: str) -> str:
    project_full_name = str(project_full_name).strip().replace("\\", "/")
    if "/" in project_full_name:
        return project_full_name.rsplit("/", 1)[-1]
    return project_full_name


def _extract_simple_method_name(raw: object) -> str:
    text = str(raw or "").strip()
    if text == "":
        return ""
    if ":" in text and "/" in text:
        text = text.rsplit(":", 1)[-1]
    if "(" in text:
        text = text.split("(", 1)[0]
    if "." in text:
        text = text.rsplit(".", 1)[-1]
    return text.strip()


def _load_benchmark_rows(
    benchmark_xlsx: Path,
    project_name: str,
    bug_type: str,
) -> List[BenchmarkBug]:
    if bug_type.upper() != "MLK":
        return []

    wb = openpyxl.load_workbook(benchmark_xlsx, read_only=True, data_only=True)
    if "details" not in wb.sheetnames:
        raise ValueError(f"'details' sheet not found in {benchmark_xlsx}")
    ws = wb["details"]

    rows: List[BenchmarkBug] = []
    for row in _iter_detail_rows(ws):
        project_full_name = str(row.get("projects", "") or "").strip()
        if project_full_name == "":
            continue
        short_name = _project_short_name(project_full_name)
        if project_name not in {project_full_name, short_name}:
            continue

        benchmark_file, benchmark_method = _parse_defect_method(row.get("defect method"))
        if benchmark_file == "":
            benchmark_file = str(row.get("file", "") or "").strip()
        rows.append(
            BenchmarkBug(
                benchmark_bug_id=str(row.get("ID", "") or ""),
                project_full_name=project_full_name,
                project_short_name=short_name,
                bug_type="MLK",
                benchmark_file=benchmark_file,
                benchmark_method=benchmark_method,
                start_line=_safe_int(row.get("start line")),
                end_line=_safe_int(row.get("end line")),
                key_variable_lines=_parse_key_variable_lines(
                    row.get("key variable location")
                ),
            )
        )
    return rows


def _extract_report_entries(
    detect_info_payload: Dict,
    context: Dict[str, str],
    result_dir: Path,
) -> List[ReportEntry]:
    entries: List[ReportEntry] = []
    for raw_report_id, item in detect_info_payload.items():
        if not isinstance(item, dict):
            continue
        metadata = item.get("metadata", {})
        if not isinstance(metadata, dict):
            metadata = {}

        relevant_functions = item.get("relevant_functions", [])
        relevant_files = []
        relevant_methods = []
        if (
            isinstance(relevant_functions, list)
            and len(relevant_functions) >= 2
            and isinstance(relevant_functions[0], list)
            and isinstance(relevant_functions[1], list)
        ):
            relevant_files = [str(v) for v in relevant_functions[0] if v is not None]
            relevant_methods = [str(v) for v in relevant_functions[1] if v is not None]

        primary_file = str(
            metadata.get("source_file")
            or (relevant_files[0] if len(relevant_files) > 0 else "")
        )
        primary_method = str(
            metadata.get("issue_primary_defect_method")
            or metadata.get("source_method_name")
            or (relevant_methods[0] if len(relevant_methods) > 0 else "")
        )
        source_line = _safe_int(metadata.get("source_line"))
        issue_source_lines = metadata.get("issue_source_lines", [])
        source_lines: List[int] = []
        if isinstance(issue_source_lines, list):
            for line in issue_source_lines:
                line_num = _safe_int(line)
                if line_num is not None:
                    source_lines.append(line_num)
        if source_line is not None and source_line not in source_lines:
            source_lines.append(source_line)
        source_lines = sorted(set(source_lines))

        issue_member_ids = metadata.get("issue_member_ids", [])
        grouped_report_ids: List[str] = []
        if isinstance(issue_member_ids, list) and len(issue_member_ids) > 0:
            grouped_report_ids = [str(v) for v in issue_member_ids]
        else:
            grouped_report_ids = [str(raw_report_id)]

        grouped_report_count = _safe_int(metadata.get("issue_member_count"))
        if grouped_report_count is None:
            grouped_report_count = len(grouped_report_ids)

        issue_component_keys = metadata.get("issue_component_keys", [])
        if not isinstance(issue_component_keys, list):
            issue_component_keys = []

        source_methods_raw = metadata.get("issue_source_methods", [])
        source_methods = []
        if isinstance(source_methods_raw, list):
            source_methods.extend(str(v) for v in source_methods_raw if v is not None)
        source_method_name = str(metadata.get("source_method_name", "") or "")
        if source_method_name != "":
            source_methods.append(source_method_name)
        if primary_method != "":
            source_methods.append(primary_method)
        source_methods = sorted(
            {
                _extract_simple_method_name(v)
                for v in source_methods
                if _extract_simple_method_name(v) != ""
            }
        )

        entries.append(
            ReportEntry(
                report_id=str(raw_report_id),
                project_name=context["project_name"],
                bug_type=context["bug_type"],
                model_name=context["model_name"],
                result_dir=str(result_dir.resolve()),
                primary_file=primary_file,
                primary_method=primary_method,
                source_line=source_line,
                source_lines=source_lines,
                issue_component_keys=[str(v) for v in issue_component_keys],
                resource_kind=str(metadata.get("resource_kind", "") or ""),
                release_context=str(metadata.get("release_context", "") or ""),
                guarantee_level=str(metadata.get("guarantee_level", "") or ""),
                grouped_report_ids=grouped_report_ids,
                grouped_report_count=grouped_report_count,
                representative_buggy_value=str(item.get("buggy_value", "") or ""),
                representative_explanation=str(item.get("explanation", "") or ""),
                source_methods=source_methods,
            )
        )
    return entries


def _paths_match(report_file: str, benchmark_file: str) -> bool:
    report_norm = _normalize_path(report_file)
    benchmark_norm = _normalize_path(benchmark_file)
    if report_norm == "" or benchmark_norm == "":
        return False
    return report_norm.endswith(benchmark_norm)


def _score_match(
    report: ReportEntry,
    benchmark: BenchmarkBug,
) -> Tuple[int, str]:
    if not _paths_match(report.primary_file, benchmark.benchmark_file):
        return 0, ""

    report_methods = {
        _extract_simple_method_name(method)
        for method in ([report.primary_method] + report.source_methods)
    }
    report_methods.discard("")
    benchmark_method = _extract_simple_method_name(benchmark.benchmark_method)
    method_match = benchmark_method != "" and benchmark_method in report_methods

    key_line_match = (
        len(benchmark.key_variable_lines) > 0
        and any(line in benchmark.key_variable_lines for line in report.source_lines)
    )

    range_match = False
    if benchmark.start_line is not None and benchmark.end_line is not None:
        range_match = any(
            benchmark.start_line <= line <= benchmark.end_line
            for line in report.source_lines
        )

    if key_line_match and method_match:
        return 100, "file+method+key_line"
    if key_line_match:
        return 90, "file+key_line"
    if method_match and range_match:
        return 70, "file+method+line_range"
    if method_match and report.source_line is not None:
        return 50, "file+method"
    return 0, ""


def _match_reports_to_benchmarks(
    reports: Sequence[ReportEntry],
    benchmarks: Sequence[BenchmarkBug],
) -> Tuple[Dict[str, str], Dict[str, str]]:
    report_to_benchmark: Dict[str, str] = {}
    report_to_rule: Dict[str, str] = {}
    for report in reports:
        best_score = -1
        best_bug_id = ""
        best_rule = ""
        for benchmark in benchmarks:
            score, rule = _score_match(report, benchmark)
            if score > best_score:
                best_score = score
                best_bug_id = benchmark.benchmark_bug_id
                best_rule = rule
        if best_score > 0 and best_bug_id != "":
            report_to_benchmark[report.report_id] = best_bug_id
            report_to_rule[report.report_id] = best_rule
    return report_to_benchmark, report_to_rule


def _benchmark_sheet_rows(
    benchmarks: Sequence[BenchmarkBug],
    reports: Sequence[ReportEntry],
    report_to_benchmark: Dict[str, str],
    report_to_rule: Dict[str, str],
    context: Dict[str, str],
    result_dir: Path,
) -> List[Dict[str, object]]:
    report_map = {report.report_id: report for report in reports}
    rows: List[Dict[str, object]] = []
    for benchmark in sorted(benchmarks, key=lambda item: (item.benchmark_bug_id, item.benchmark_file)):
        matched_reports = [
            report_map[report_id]
            for report_id, bug_id in report_to_benchmark.items()
            if bug_id == benchmark.benchmark_bug_id and report_id in report_map
        ]
        matched_report_ids = [report.report_id for report in matched_reports]
        matched_buggy_values = [
            report.representative_buggy_value for report in matched_reports
        ]
        matched_rules = sorted(
            {
                report_to_rule.get(report.report_id, "")
                for report in matched_reports
                if report_to_rule.get(report.report_id, "") != ""
            }
        )
        rows.append(
            {
                "benchmark_bug_id": benchmark.benchmark_bug_id,
                "project_name": context["project_name"],
                "bug_type": context["bug_type"],
                "model_name": context["model_name"],
                "result_dir": str(result_dir.resolve()),
                "benchmark_file": benchmark.benchmark_file,
                "benchmark_line": (
                    ",".join(str(v) for v in benchmark.key_variable_lines)
                    if len(benchmark.key_variable_lines) > 0
                    else (
                        f"{benchmark.start_line}-{benchmark.end_line}"
                        if benchmark.start_line is not None
                        and benchmark.end_line is not None
                        else ""
                    )
                ),
                "is_detected": "yes" if len(matched_reports) > 0 else "no",
                "matched_report_count": len(matched_reports),
                "matched_report_ids": _json_cell(matched_report_ids),
                "matched_buggy_values": _json_cell(matched_buggy_values),
                "match_rule": "; ".join(matched_rules),
                "final_label": "TP_known" if len(matched_reports) > 0 else "FN",
                "note": "",
            }
        )
    return rows


def _review_unit_rows(
    reports: Sequence[ReportEntry],
    report_to_benchmark: Dict[str, str],
) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    unmatched_reports = [
        report for report in reports if report.report_id not in report_to_benchmark
    ]
    unmatched_reports = sorted(
        unmatched_reports,
        key=lambda item: (
            _normalize_path(item.primary_file),
            item.source_line if item.source_line is not None else -1,
            item.report_id,
        ),
    )
    for report in unmatched_reports:
        rows.append(
            {
                "review_unit_id": f"RU-{report.report_id}",
                "project_name": report.project_name,
                "bug_type": report.bug_type,
                "model_name": report.model_name,
                "result_dir": report.result_dir,
                "primary_file": report.primary_file,
                "primary_method": report.primary_method,
                "source_line": report.source_line if report.source_line is not None else "",
                "resource_kind": report.resource_kind,
                "release_context": report.release_context,
                "guarantee_level": report.guarantee_level,
                "representative_report_id": report.report_id,
                "grouped_report_count": report.grouped_report_count,
                "grouped_report_ids": _json_cell(report.grouped_report_ids),
                "representative_buggy_value": report.representative_buggy_value,
                "representative_explanation": report.representative_explanation,
                "benchmark_relation": "non_benchmark_candidate",
                "review_status": "todo",
                "human_label": "",
                "decision_note": "",
            }
        )
    return rows


def _write_sheet(ws, headers: Sequence[str], rows: Sequence[Dict[str, object]]) -> None:
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wrap_columns = {
        "matched_buggy_values",
        "matched_report_ids",
        "grouped_report_ids",
        "representative_buggy_value",
        "representative_explanation",
        "decision_note",
        "note",
    }

    for idx, header in enumerate(headers, start=1):
        column_letter = openpyxl.utils.get_column_letter(idx)
        if header in wrap_columns:
            width = 36
        elif header.endswith("_id") or header.endswith("_line"):
            width = 16
        elif "file" in header or "dir" in header:
            width = 40
        else:
            width = 18
        ws.column_dimensions[column_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_review_workbook(
    result_dir: Path,
    benchmark_xlsx: Path,
    output_path: Optional[Path] = None,
) -> Path:
    result_dir = result_dir.resolve()
    if output_path is None:
        output_path = result_dir / "review_units.xlsx"
    else:
        output_path = output_path.resolve()

    context = _parse_result_dir_context(result_dir)
    detect_info = _load_json(result_dir / "detect_info.json")
    benchmarks = _load_benchmark_rows(
        benchmark_xlsx=benchmark_xlsx,
        project_name=context["project_name"],
        bug_type=context["bug_type"],
    )
    reports = _extract_report_entries(detect_info, context=context, result_dir=result_dir)
    report_to_benchmark, report_to_rule = _match_reports_to_benchmarks(reports, benchmarks)

    benchmark_rows = _benchmark_sheet_rows(
        benchmarks=benchmarks,
        reports=reports,
        report_to_benchmark=report_to_benchmark,
        report_to_rule=report_to_rule,
        context=context,
        result_dir=result_dir,
    )
    review_rows = _review_unit_rows(reports=reports, report_to_benchmark=report_to_benchmark)

    wb = openpyxl.Workbook()
    benchmark_ws = wb.active
    benchmark_ws.title = "benchmark_hits"
    review_ws = wb.create_sheet("review_units")

    benchmark_headers = [
        "benchmark_bug_id",
        "project_name",
        "bug_type",
        "model_name",
        "result_dir",
        "benchmark_file",
        "benchmark_line",
        "is_detected",
        "matched_report_count",
        "matched_report_ids",
        "matched_buggy_values",
        "match_rule",
        "final_label",
        "note",
    ]
    review_headers = [
        "review_unit_id",
        "project_name",
        "bug_type",
        "model_name",
        "result_dir",
        "primary_file",
        "primary_method",
        "source_line",
        "resource_kind",
        "release_context",
        "guarantee_level",
        "representative_report_id",
        "grouped_report_count",
        "grouped_report_ids",
        "representative_buggy_value",
        "representative_explanation",
        "benchmark_relation",
        "review_status",
        "human_label",
        "decision_note",
    ]

    _write_sheet(benchmark_ws, benchmark_headers, benchmark_rows)
    _write_sheet(review_ws, review_headers, review_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build review_units.xlsx for one RepoAudit result directory."
    )
    parser.add_argument(
        "--result-dir",
        required=True,
        help="One run result directory, e.g. result/.../<timestamp>",
    )
    parser.add_argument(
        "--benchmark-xlsx",
        default=str(DEFAULT_BENCHMARK_XLSX),
        help="Path to selected_real_world_projects.xlsx",
    )
    parser.add_argument(
        "--output",
        default="",
        help="Optional output xlsx path (default: <result-dir>/review_units.xlsx)",
    )
    args = parser.parse_args()

    result_dir = Path(args.result_dir)
    benchmark_xlsx = Path(args.benchmark_xlsx)
    output_path = Path(args.output) if str(args.output).strip() != "" else None

    output = build_review_workbook(
        result_dir=result_dir,
        benchmark_xlsx=benchmark_xlsx,
        output_path=output_path,
    )
    print(f"[OK] review workbook written to: {output}")


if __name__ == "__main__":
    main()
